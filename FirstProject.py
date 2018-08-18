from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active
wb_read = load_workbook(filename='Data.xlsx')
ws_read_orders = wb_read['Orders']
ws_read_inventory = wb_read['Inventory']

print(wb_read)

total_inventory = []
total_orders = []
i = 0
for row in ws_read_orders.rows:
    a = []
    if i != 0:
        a.extend([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value ])
        total_orders.append(a)
    i = 1

print(total_orders)

tickers = []
for i in total_inventory:
    if i[1] not in tickers:
        tickers.append(i[1])
for i in total_orders:
    if i[1] not in tickers:
        tickers.append(i[1])
all_data = []
for ticker in tickers:
    append_inv = []
    append_order = []
    for i in total_inventory:
        if i[1] == ticker:
            append_inv.append(i)
    for i in total_orders:
        if i[1] == ticker:
            append_order.append(i)
    a = {   'ticker' : ticker,
            'inventory' : append_inv,
            'orders': append_order
    }
    all_data.append(a)


    # start the FIFO method
for data in all_data:
    text = "Thống kê cho mã chứng khoán" + data['ticker']
    ws.append([text])
    ws.append(['Thời gian', 'Số hiệu lệnh','Loại giao dịch','Nội dung','Lãi lỗ thực hiện', 'Nợ/ có', 'Phí / thuế', 'Lãi tiền gửi/ lãi vay', 'Số dư tài khoản'])

    # setupdata
    inventory = data['inventory']
    orders = data['orders']

    for order in orders:
        if not len(inventory) != 0 or order[3] == inventory[0][3]:
            inventory.append(order)
            order.extend(['inv stack',0])
            ws.append(order)
        else:
            prices = []
            total_inv = 0
            prices_if = []
            for inv in inventory:
                total_inv += inv[4]
                prices_if.append(inv[5])
            if total_inv < order[4]:
                ws.append([order[0], order[1], order[2], order[3], order[4] - total_inv, order[5], "inv stack", 0 ])
                if order[3] == "Mua":
                    profit = -(total_inv)*(order[4] - sum(prices_if)/len(prices_if))
                else:
                    profit = (total_inv)*(order[4] - sum(prices_if)/len(prices_if))
                ws.append([ order[0], order[1], order[2], order[3], total_inv, order[5], sum(prices_if)/len(prices_if), profit ])
            else:
                for i in range (0, order[4]):
                    prices.append(inventory[0][5])
                    inventory[0][4] -= 1
                    if inventory[0][4] == 0:
                        inventory.pop(0)
                if order[3] == "Mua":
                    profit = (order[4])*(order[5] - sum(prices)/len(prices))
                else:
                    profit = -(order[4])*(order[5] - sum(prices)/len(prices))
                order.extend([sum(prices)/len(prices), profit  ])
                ws.append(order)


wb.save("Result.xlsx")







